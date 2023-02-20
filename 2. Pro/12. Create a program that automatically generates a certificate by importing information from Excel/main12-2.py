#openpyxl을 이용해 엑셀 불러오기.
#값을 읽고 출력.

from openpyxl import load_workbook

load_wb = load_workbook(r"40 Python works\12. Create a program that automatically generates a certificate by importing information from Excel\수료증명단.xlsx")
load_ws = load_wb.active

name_list = [] #빈 리스트 3개 생성
birthday_list = [] 
ho_list = []

for i in range(1, load_ws.max_row + 1): #엑셀파일의 마지막줄까지 줄을 읽고 어펜드로 리스트값들 추가해줌.
#3가지 리스트 만들어 각 리스트 값들을 어펜드 해주고 3가지 리스트를 출력.
    name_list.append(load_ws.cell(i, 1). value)
    birthday_list.append(load_ws.cell(i, 2).value)
    ho_list.append(load_ws.cell(i, 3).value)
#인터넷에서 찾은 새 방법
#doc = Document(r'경로\수료증.docx)
#for i, paragraph in enumerate(doc.paragraphs):
    #print(str(i) + ": " + paragraph.text)

print(name_list)
print(birthday_list)
print(ho_list)